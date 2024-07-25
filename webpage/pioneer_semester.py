import openpyxl as op
from flask import Flask, render_template
from flask_cors import CORS
import os

app = Flask(__name__)
CORS(app)

@app.route('/')
def home():
    try:
        # 첫 번째 엑셀 파일 로드
        file1 = r"C:\Users\USER\Desktop\비교과\flask\.vscode\detect_result.xlsx"
        if not os.path.exists(file1):
            return f"File not found: {file1}"
        
        wb = op.load_workbook(file1)
        ws = wb.active

        car_images = [] 
        plate_texts = [] 
        plate_images = []
        latitude_degs = []
        longitude_degs = []
        capture_times = []

        for row in ws.iter_rows(min_row=2, values_only=True):  # 첫 번째 행은 헤더로 가정
            plate_text = row[2]
            latitude_deg = row[4]
            longitude_deg = row[5]
            capture_time = row[3]
            if plate_text is None or latitude_deg is None or longitude_deg is None or capture_time is None:
                continue  # 텍스트가 존재하지 않는 행을 만나면 반복 종료
            car_images.append(f"https://detectpath.s3.ap-northeast-2.amazonaws.com/{capture_time}.webp")
            plate_images.append(f"https://detectpath.s3.ap-northeast-2.amazonaws.com/{capture_time}plate.webp")
            plate_texts.append(plate_text)
            latitude_degs.append(latitude_deg)
            longitude_degs.append(longitude_deg)
            capture_times.append(capture_time)
        
        # 두 번째 엑셀 파일 로드
        file2 = r"C:\Users\USER\Desktop\비교과\flask\.vscode\noinformation.xlsx"
        if not os.path.exists(file2):
            return f"File not found: {file2}"

        wb2 = op.load_workbook(file2)
        ws2 = wb2.active

        parking_alts = []
        parking_longs = []

        for row in ws2.iter_rows(min_row=2, values_only=True):  # 첫 번째 행은 헤더로 가정
           
            parking_alt = row[0]
            parking_long = row[1]
            if parking_alt is None or parking_long is None:
                continue  # 텍스트가 존재하지 않는 행을 만나면 반복 종료
            parking_alts.append(parking_alt)
            parking_longs.append(parking_long)

        # 디버깅 출력
        print("Car Images:", car_images)
        print("Plate Texts:", plate_texts)
        print("Latitude Degrees:", latitude_degs)
        print("Longitude Degrees:", longitude_degs)
        print("Capture Times:", capture_times)
        print("Parking Alts:", parking_alts)
        print("Parking Longs:", parking_longs)

        return render_template(
            'parking.html',
            car_images=car_images,
            plate_texts=plate_texts,
            plate_images=plate_images,
            latitude_degs=latitude_degs,
            longitude_degs=longitude_degs,
            capture_times=capture_times,
            parking_alts=parking_alts,
            parking_longs=parking_longs
        )
    except Exception as e:
        return str(e)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)

