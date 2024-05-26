import openpyxl as op
from flask import Flask, render_template
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

@app.route('/')
def home():
    # 엑셀 파일에서 데이터 읽어오기
    wb = op.load_workbook(r"C:\\Users\\USER\\Desktop\\data.xlsx")
    ws = wb.active

    # 데이터 추출
    car_image = "https://detectpath.s3.ap-northeast-2.amazonaws.com/2024-05-20-12081111.webp"
    plate_text = ws["B1"].value
    plate_image = "https://detectpath.s3.ap-northeast-2.amazonaws.com/2024-05-20-12081111plate.webp"
    latitude_deg = ws["D1"].value
    longitude_deg = ws["E1"].value
    capture_time = ws["F1"].value

    # render_template 함수를 호출할 때 변수를 함께 전달
    return render_template('space.html', car_image=car_image, plate_text=plate_text, plate_image=plate_image, latitude_deg=latitude_deg, longitude_deg=longitude_deg, capture_time=capture_time)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=80)  # 포트 80 사용

